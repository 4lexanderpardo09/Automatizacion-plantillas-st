# logic.py

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io

# --- DICCIONARIOS Y CONSTANTES ---

REGIONES = {
    'BORDO': ['10201', '10204'],
    'POPAYAN': ['10101'],
    'SANTANDER': ['103', '10301'],
    'AMBIENTA': ['10401'],
    'VALLE': ['20101'],
    'PASTO': ['30101'],
    'TUQUERRES': ['30301'],
    'PITALITO': ['70101', '70102', '70104']
}

MAPEO_COLUMNAS = {
    'REGION': 'REGION',
    'C DE COSTO': 'CCOSER',
    'O. DE SERVICIO': 'NUM_OS',
    'DIAS': 'DIAS',
    'CARTA ABANDONO': 'CARTA ABANDONO',
    'HOY': 'HOY',
    'FECHA INGRESO': 'FECHA',
    'NOMBRE': 'CLIENTE',
    'CEDULA': 'CEDULA',
    'OBSERVACION': 'DETALLE_AC',
    'ARTICULO': 'PRODUCTO',
    'ESTADO': 'DETALLE',
    'SERIE': 'SERIE',
    'FALLA': 'CONCEPTO_E'
}

# Definición de colores
NARANJA = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
VERDE = PatternFill(start_color='00BB2D', end_color='00BB2D', fill_type='solid')
AMARILLO = PatternFill(start_color='BFFF00', end_color='BFFF00', fill_type='solid')
ROJO = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
GRIS = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')


# --- FUNCIONES DE PROCESAMIENTO ---

def aplicar_colores_semaforo(ws):
    """Aplica el relleno de color a la columna 'DIAS' de una hoja de cálculo."""
    col_dias_letra = None
    col_carta_letra = None

    for col in ws.iter_cols(min_row=1, max_row=1):
        if col[0].value == 'DIAS':
            col_dias_letra = col[0].column_letter
        if col[0].value == 'CARTA ABANDONO':
            col_carta_letra = col[0].column_letter

    if not col_dias_letra or not col_carta_letra:
        print(f"ADVERTENCIA: Columnas 'DIAS' o 'CARTA ABANDONO' no encontradas en la hoja '{ws.title}'.")
        return

    for row in range(2, ws.max_row + 1):
        celda_carta = ws[f"{col_carta_letra}{row}"]
        celda_dias = ws[f"{col_dias_letra}{row}"]
        
        if str(celda_carta.value).strip().lower() == "si":
            celda_dias.fill = NARANJA
            continue
            
        try:
            dias = int(celda_dias.value)
            if 1 <= dias <= 10: celda_dias.fill = VERDE
            elif 11 <= dias <= 20: celda_dias.fill = AMARILLO
            elif 21 <= dias <= 31: celda_dias.fill = ROJO
            elif dias >= 32: celda_dias.fill = GRIS
        except (ValueError, TypeError):
            continue

def generar_reporte_region(df_region, nombre_region):
    """
    Toma un DataFrame de una región, calcula los campos necesarios,
    mapea las columnas y devuelve el archivo Excel con formato en bytes.
    """
    # 1. Calcular columnas nuevas
    df = df_region.copy()
    df['FECHA'] = pd.to_datetime(df['FECHA'], errors='coerce').dt.date
    df['HOY'] = datetime.now().date()
    df['DIAS'] = df.apply(lambda row: (row['HOY'] - row['FECHA']).days if pd.notnull(row['FECHA']) else None, axis=1)
    df['CARTA ABANDONO'] = df['DETALLE_AC'].fillna('').str.lower().str.contains('carta abandono').map({True: 'Si', False: 'No'})

    # 2. Crear el DataFrame final con las columnas correctas
    df_final = pd.DataFrame()
    for col_final, col_origen in MAPEO_COLUMNAS.items():
        df_final[col_final] = df.get(col_origen)

    # 3. Crear el archivo Excel en un buffer de memoria
    output_buffer = io.BytesIO()
    df_final.to_excel(output_buffer, index=False, engine='openpyxl')
    
    # 4. Aplicar formato (colores y ancho de columnas)
    output_buffer.seek(0)
    wb = load_workbook(output_buffer)
    ws = wb.active

    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value))
        ws.column_dimensions[col[0].column_letter].width = max_len + 2
        
    aplicar_colores_semaforo(ws)
    
    # 5. Guardar los cambios finales en un nuevo buffer y devolverlo
    final_buffer = io.BytesIO()
    wb.save(final_buffer)
    
    return final_buffer.getvalue()

def cargar_y_filtrar_datos(archivo_cargado):
    """Carga el archivo, lo limpia y lo filtra, devolviendo un DataFrame listo para procesar."""
    df_informe = pd.read_excel(archivo_cargado)
    df_informe.columns = df_informe.columns.str.strip().str.upper()

    if 'CCOSER' not in df_informe.columns or 'ESTADONOMB' not in df_informe.columns:
        raise ValueError("El archivo debe contener las columnas 'CCOSER' y 'ESTADONOMB'.")
        
    df_informe['CCOSER'] = df_informe['CCOSER'].astype(str).str.strip()
    df_informe['REGION'] = None
    
    for region, codigos in REGIONES.items():
        df_informe.loc[df_informe['CCOSER'].isin(codigos), 'REGION'] = region
        
    df_pendientes = df_informe[df_informe['ESTADONOMB'].str.strip().str.upper() == 'PENDIENTE'].copy()
    return df_pendientes